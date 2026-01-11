#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import os
import tempfile
import datetime
import shutil
import webbrowser
import threading
import requests
import json

# 导入核心比较函数
from compare_excel_web import compare_excel_files

# 初始化FastAPI应用
app = FastAPI(
    title="Excel比较工具API",
    description="高效的Excel文件比较服务",
    version="1.0.0"
)

# 添加CORS中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 定义结果文件夹
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
# 在Vercel上，只有/tmp目录是可写的，所以使用/tmp/results
RESULTS_FOLDER = os.path.join("/tmp", "results")
os.makedirs(RESULTS_FOLDER, exist_ok=True)

# 挂载静态文件到/static路径
app.mount("/static", StaticFiles(directory=PROJECT_ROOT), name="static")

# 提供主页面路由
@app.get("/")
async def root():
    return FileResponse(os.path.join(PROJECT_ROOT, "index.html"))

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """下载结果文件"""
    try:
        # 构建完整的文件路径
        file_path = os.path.join(RESULTS_FOLDER, filename)
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")
        
        # 返回文件下载响应
        return FileResponse(file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/get_project_info")
async def get_project_info():
    """获取项目信息，包括最新版本和最后更新日期"""
    try:
        # Gitee API参数
        owner = "caifugao110"
        repo = "table-comparison-hyl"
        gitee_token = "a09da64c1d9e9c7420a18dfd838890b0"
        headers = {
            "Authorization": f"token {gitee_token}",
            "Accept": "application/json"
        }
        
        # 初始化变量
        version = None
        last_update_date = None
        
        # 获取最新发行版本
        try:
            release_url = f"https://gitee.com/api/v5/repos/{owner}/{repo}/releases/latest"
            release_response = requests.get(release_url, headers=headers, timeout=10)
            if release_response.status_code == 200:
                release_data = release_response.json()
                version = release_data.get("tag_name")
        except Exception as e:
            print(f"获取最新发行版本失败: {e}")
        
        # 获取最后一次提交日期
        try:
            commit_url = f"https://gitee.com/api/v5/repos/{owner}/{repo}/commits?per_page=1"
            commit_response = requests.get(commit_url, headers=headers, timeout=10)
            if commit_response.status_code == 200:
                commit_data = commit_response.json()
                if isinstance(commit_data, list) and len(commit_data) > 0:
                    commit_date_str = commit_data[0]["commit"]["committer"]["date"]
                    commit_date = datetime.datetime.fromisoformat(commit_date_str.replace("Z", ""))
                    last_update_date = f"{commit_date.year}年{commit_date.month:02d}月"
        except Exception as e:
            print(f"获取最后提交日期失败: {e}")
        
        # 返回结果，不使用默认值
        return JSONResponse({
            "version": version,
            "lastUpdateDate": last_update_date
        })
    except Exception as e:
        # 记录错误但不使用默认值
        print(f"获取项目信息失败: {e}")
        return JSONResponse({
            "version": None,
            "lastUpdateDate": None
        })

@app.post("/api/preview")
async def preview_excel(
    baselineFile: UploadFile = File(...),
    header_row: int = Form(...)
):
    """预览Excel文件的表头行和特征列"""
    try:
        import openpyxl
        
        # 保存上传的文件到临时位置
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_baseline:
            temp_baseline.write(await baselineFile.read())
            baseline_file_path = temp_baseline.name
        
        try:
            # 加载Excel文件
            wb = openpyxl.load_workbook(baseline_file_path, data_only=True)
            ws = wb.active
            
            # 获取前10行数据用于预览
            max_preview_row = min(10, ws.max_row)
            max_preview_col = min(20, ws.max_column)
            
            # 预览数据
            preview_data = []
            for r in range(1, max_preview_row + 1):
                row_data = []
                for c in range(1, max_preview_col + 1):
                    cell_value = ws.cell(row=r, column=c).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                preview_data.append({
                    "row": r,
                    "data": row_data
                })
            
            # 获取表头行数据
            header_data = []
            if header_row <= ws.max_row:
                for c in range(1, max_preview_col + 1):
                    cell_value = ws.cell(row=header_row, column=c).value
                    header_data.append({
                        "col": c,
                        "name": str(cell_value) if cell_value is not None else f"列{c}"
                    })
            
            return JSONResponse({
                "success": True,
                "preview_data": preview_data,
                "header_data": header_data,
                "max_row": ws.max_row,
                "max_col": ws.max_column
            })
        finally:
            # 清理临时文件
            os.unlink(baseline_file_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/compare")
async def compare_excel(
    baselineFile: UploadFile = File(...),
    compareFile: UploadFile = File(...),
    header_row: int = 3,
    key_fields: str = None
):
    """比较两个Excel文件"""
    try:
        # 处理特征列参数
        parsed_key_fields = None
        if key_fields:
            try:
                # 解析JSON格式的特征列
                parsed_key_fields = json.loads(key_fields)
            except json.JSONDecodeError:
                # 如果不是JSON格式，尝试解析为逗号分隔的字符串
                parsed_key_fields = [field.strip() for field in key_fields.split(",") if field.strip()]
        
        # 生成唯一的文件名和时间戳
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        original_filename = os.path.splitext(baselineFile.filename)[0]
        
        # 保存上传的文件到临时位置
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_baseline:
            temp_baseline.write(await baselineFile.read())
            baseline_file_path = temp_baseline.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_compare:
            temp_compare.write(await compareFile.read())
            compare_file_path = temp_compare.name
        
        # 生成结果文件路径
        result_baseline = os.path.join(RESULTS_FOLDER, f"{original_filename}_my_比较结果_{timestamp}.xlsx")
        result_compare = os.path.join(RESULTS_FOLDER, f"{original_filename}_from_比较结果_{timestamp}.xlsx")
        
        # 直接调用核心比较函数，使用多进程执行以提高性能
        import io
        import sys
        from contextlib import redirect_stdout
        
        f = io.StringIO()
        with redirect_stdout(f):
            compare_excel_files(
                baseline_file_path,  # 基准文件路径
                compare_file_path,   # 比较文件路径
                result_baseline,     # 输出基准文件路径
                result_compare,      # 输出比较文件路径
                original_filename,   # 原始文件名
                timestamp,           # 时间戳
                header_row,          # 表头行号
                parsed_key_fields    # 特征列
            )
        
        # 获取函数输出
        stdout = f.getvalue()
        
        # 生成差异结果文件路径
        diff_file = os.path.join(RESULTS_FOLDER, f"{original_filename}_差异结果_{timestamp}.xlsx")
        
        # 收集所有结果文件，只返回文件名
        result_files = []
        expected_files = [diff_file, result_baseline, result_compare]
        for expected_file in expected_files:
            if os.path.exists(expected_file):
                # 只返回文件名，不返回完整路径
                result_files.append(os.path.basename(expected_file))
        
        # 清理临时文件
        os.unlink(baseline_file_path)
        os.unlink(compare_file_path)
        
        # 返回结果
        return JSONResponse({
            "success": True,
            "message": "比较完成",
            "resultFiles": result_files,
            "stdout": stdout,
            "stderr": ""
        })
        
    except Exception as e:
        # 清理临时文件
        if 'baseline_file_path' in locals() and os.path.exists(baseline_file_path):
            os.unlink(baseline_file_path)
        if 'compare_file_path' in locals() and os.path.exists(compare_file_path):
            os.unlink(compare_file_path)
        
        raise HTTPException(status_code=500, detail=str(e))

def open_browser():
    """延迟打开浏览器，确保服务器已经启动"""
    import time
    time.sleep(2)  # 延迟2秒，等待服务器启动
    webbrowser.open("http://localhost:8000")

if __name__ == "__main__":
    import uvicorn
    # 在单独的线程中打开浏览器
    threading.Thread(target=open_browser, daemon=True).start()
    # 启动服务器
    uvicorn.run(app, host="0.0.0.0", port=8000)
