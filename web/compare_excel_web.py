import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess
import stat


def compare_excel_files(baseline_path, compare_path, output_baseline_path, output_compare_path, original_filename, timestamp, header_row=3, key_fields=None):
    # 获取文件夹名称用于标识
    baseline_folder = os.path.basename(os.path.dirname(baseline_path))
    compare_folder = os.path.basename(os.path.dirname(compare_path))
    
    # 定义颜色样式
    fill_changed = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色：数值变化
    fill_added = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")      # 绿色：新增（在基准基础上）
    fill_deleted = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # 红色：删除（在基准基础上）

    print(f"正在加载文件: {baseline_path} 和 {compare_path} ...")
    
    try:
        # 加载工作簿
        wb_baseline = openpyxl.load_workbook(baseline_path, data_only=True)  # 只加载数据，不加载公式
        wb_compare = openpyxl.load_workbook(compare_path, data_only=True)
    except FileNotFoundError as e:
        print(f"错误：找不到文件 - {e}")
        return
    except Exception as e:
        print(f"加载文件时出错: {e}")
        return

    # 1. 选择工作表
    print(f"\n【{baseline_folder}文件夹】工作表列表: {wb_baseline.sheetnames}")
    print(f"【{compare_folder}文件夹】工作表列表: {wb_compare.sheetnames}")
    
    # 默认使用第一个工作表
    ws_baseline = wb_baseline.active
    ws_compare = wb_compare.active
    print(f"\n默认比较第一个工作表: {ws_baseline.title} ({baseline_folder}) vs {ws_compare.title} ({compare_folder})")

    # 2. 获取实际使用的范围
    baseline_max_row = ws_baseline.max_row
    baseline_max_col = ws_baseline.max_column
    compare_max_row = ws_compare.max_row
    compare_max_col = ws_compare.max_column

    print(f"开始比较 ({baseline_folder}文件夹: {baseline_max_row}行 x {baseline_max_col}列, {compare_folder}文件夹: {compare_max_row}行 x {compare_max_col}列)...")

    # 3. 预先获取所有单元格值
    cells_baseline = {}
    cells_compare = {}
    
    # 获取基准文件所有单元格值
    for r in range(1, baseline_max_row + 1):
        for c in range(1, baseline_max_col + 1):
            cells_baseline[(r, c)] = ws_baseline.cell(row=r, column=c).value
    
    # 获取比较文件所有单元格值
    for r in range(1, compare_max_row + 1):
        for c in range(1, compare_max_col + 1):
            cells_compare[(r, c)] = ws_compare.cell(row=r, column=c).value
    
    # 4. 基于关键字段的行匹配算法
    def get_col_content(col_num, cells, max_row):
        """获取一列的所有单元格内容，作为比较的键"""
        return tuple(cells.get((r, col_num), None) for r in range(1, max_row + 1))
    
    # 如果没有提供关键字段，使用默认关键字段
    if not key_fields:
        key_fields = ["部门", "合同号", "产品代码"]
    
    # 从指定行获取关键字段的列索引
    def find_key_columns(ws, cells, max_col, header_row_num):
        """从指定行查找关键字段的列索引"""
        key_cols = {}
        for col in range(1, max_col + 1):
            cell_value = cells.get((header_row_num, col), "").strip()  # 第header_row行是表头
            if cell_value in key_fields:
                key_cols[cell_value] = col
        return key_cols
    
    # 查找基准文件和比较文件的关键字段列索引
    key_cols_baseline = find_key_columns(ws_baseline, cells_baseline, baseline_max_col, header_row)
    key_cols_compare = find_key_columns(ws_compare, cells_compare, compare_max_col, header_row)
    
    print(f"\n基准文件关键字段列索引: {key_cols_baseline}")
    print(f"比较文件关键字段列索引: {key_cols_compare}")
    
    # 检查是否找到所有关键字段
    has_all_keys_baseline = all(field in key_cols_baseline for field in key_fields)
    has_all_keys_compare = all(field in key_cols_compare for field in key_fields)
    
    # 行匹配：基准行号 -> 比较行号
    row_mapping = {}
    
    if has_all_keys_baseline and has_all_keys_compare:
        print("\n使用关键字段进行行匹配...")
        
        # 构建行关键字映射：关键字 -> 行号
        def build_row_key_map(cells, max_row, key_cols, data_start_row):
            row_key_map = {}
            for row in range(data_start_row, max_row + 1):  # 从数据行开始
                key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
                # 只有当所有关键字段都有值时才进行映射
                if all(v is not None for v in key_values):
                    row_key_map[key_values] = row
            return row_key_map
        
        # 数据行从表头行的下一行开始
        data_start_row = header_row + 1
        row_key_map_baseline = build_row_key_map(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        row_key_map_compare = build_row_key_map(cells_compare, compare_max_row, key_cols_compare, data_start_row)
        
        # 建立行映射：基准行 -> 比较行
        for key in row_key_map_baseline:
            if key in row_key_map_compare:
                row_baseline = row_key_map_baseline[key]
                row_compare = row_key_map_compare[key]
                row_mapping[row_baseline] = row_compare
        
        print(f"基于关键字段匹配到 {len(row_mapping)} 行")
    else:
        print("\n无法找到所有关键字段，使用默认行匹配...")
        # 原来的行匹配逻辑
        def get_row_content(row_num, cells, max_col):
            """获取一行的所有单元格内容，作为比较的键"""
            return tuple(cells.get((row_num, c), None) for c in range(1, max_col + 1))
        
        # 构建行内容映射
        row_contents_baseline = {r: get_row_content(r, cells_baseline, baseline_max_col) for r in range(1, baseline_max_row + 1)}
        row_contents_compare = {r: get_row_content(r, cells_compare, compare_max_col) for r in range(1, compare_max_row + 1)}
        
        # 先找到完全匹配的行
        for row_baseline in row_contents_baseline:
            content_baseline = row_contents_baseline[row_baseline]
            for row_compare in row_contents_compare:
                if row_compare not in row_mapping.values() and content_baseline == row_contents_compare[row_compare]:
                    row_mapping[row_baseline] = row_compare
                    break
        
        # 如果没有找到足够的匹配，使用简单的索引映射
        if len(row_mapping) < min(baseline_max_row, compare_max_row) // 2:
            min_rows = min(baseline_max_row, compare_max_row)
            row_mapping = {r: r for r in range(1, min_rows + 1)}
    
    # 列匹配：基准列号 -> 比较列号
    col_mapping = {}
    
    # 构建列内容映射
    col_contents_baseline = {c: get_col_content(c, cells_baseline, baseline_max_row) for c in range(1, baseline_max_col + 1)}
    col_contents_compare = {c: get_col_content(c, cells_compare, compare_max_row) for c in range(1, compare_max_col + 1)}
    
    # 先找到完全匹配的列
    for col_baseline in col_contents_baseline:
        content_baseline = col_contents_baseline[col_baseline]
        for col_compare in col_contents_compare:
            if col_compare not in col_mapping.values() and content_baseline == col_contents_compare[col_compare]:
                col_mapping[col_baseline] = col_compare
                break
    
    # 如果没有找到足够的匹配，使用简单的索引映射
    if len(col_mapping) < min(baseline_max_col, compare_max_col) // 2:
        min_cols = min(baseline_max_col, compare_max_col)
        col_mapping = {c: c for c in range(1, min_cols + 1)}
    
    # 5. 比较单元格
    changes_count = 0
    
    # 定义关键字段列索引集合，避免重新计算
    key_col_set_baseline = set(key_cols_baseline.values()) if has_all_keys_baseline else set()
    key_col_set_compare = set(key_cols_compare.values()) if has_all_keys_compare else set()
    
    # 只比较匹配的行（基于关键字段匹配的行）
    print("\n开始比较匹配行的单元格差异...")
    
    for row_baseline in row_mapping:
        row_compare = row_mapping[row_baseline]
        
        # 获取当前行的所有列索引
        baseline_cols = range(1, baseline_max_col + 1)
        compare_cols = range(1, compare_max_col + 1)
        
        # 创建列映射（基于列名匹配）
        col_name_map = {}
        for col_b in baseline_cols:
            # 获取基准文件列名
            col_name_b = cells_baseline.get((header_row, col_b), "").strip()
            if not col_name_b:
                continue
            
            # 在比较文件中查找相同列名
            for col_c in compare_cols:
                col_name_c = cells_compare.get((header_row, col_c), "").strip()
                if col_name_c == col_name_b:
                    col_name_map[col_b] = col_c
                    break
        
        # print(f"行 {row_baseline} -> {row_compare} 的列映射: {col_name_map}")
        
        # 比较匹配的列
        for col_baseline, col_compare in col_name_map.items():
            # 跳过关键字段列（它们已经匹配，不需要比较）
            if col_baseline in key_col_set_baseline or col_compare in key_col_set_compare:
                continue
            
            val_baseline = cells_baseline.get((row_baseline, col_baseline), None)
            val_compare = cells_compare.get((row_compare, col_compare), None)
            
            # 只在值不同时标记为黄色（数值变化）
            if val_baseline != val_compare:
                ws_baseline.cell(row=row_baseline, column=col_baseline).fill = fill_changed
                ws_compare.cell(row=row_compare, column=col_compare).fill = fill_changed
                changes_count += 1
    
    # 6. 标记新增行和删除行
    print("\n开始标记新增行和删除行...")
    
    # 获取所有数据行的关键字映射
    def get_all_row_keys(cells, max_row, key_cols, data_start_row):
        """获取所有数据行的关键字映射"""
        all_row_keys = {}
        for row in range(data_start_row, max_row + 1):  # 从数据行开始
            key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
            if all(v is not None for v in key_values):
                all_row_keys[key_values] = row
        return all_row_keys
    
    if has_all_keys_baseline and has_all_keys_compare:
        # 获取所有数据行的关键字映射
        data_start_row = header_row + 1
        all_baseline_keys = get_all_row_keys(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        all_compare_keys = get_all_row_keys(cells_compare, compare_max_row, key_cols_compare, data_start_row)
        
        # 标记删除行（基准文件中有，比较文件中没有）
        deleted_rows = 0
        for key, row_baseline in all_baseline_keys.items():
            if key not in all_compare_keys:
                # 标记整行为绿色
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                changes_count += 1
                deleted_rows += 1
        print(f"已标记 {deleted_rows} 行删除（绿色）")
        
        # 标记新增行（比较文件中有，基准文件中没有）
        added_rows = 0
        for key, row_compare in all_compare_keys.items():
            if key not in all_baseline_keys:
                # 标记整行为红色
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                changes_count += 1
                added_rows += 1
        print(f"已标记 {added_rows} 行新增（红色）")
    else:
        # 使用简单的行匹配来标记新增和删除行
        print("\n使用简单匹配标记新增和删除行...")
        
        # 标记删除行（基准文件中有，比较文件中没有对应的行）
        deleted_rows = 0
        for row_baseline in range(1, baseline_max_row + 1):
            if row_baseline not in row_mapping:
                # 标记整行为绿色
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                changes_count += 1
                deleted_rows += 1
        print(f"已标记 {deleted_rows} 行删除（绿色）")
        
        # 标记新增行（比较文件中有，基准文件中没有对应的行）
        added_rows = 0
        mapped_compare_rows = set(row_mapping.values())
        for row_compare in range(1, compare_max_row + 1):
            if row_compare not in mapped_compare_rows:
                # 标记整行为红色
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                changes_count += 1
                added_rows += 1
        print(f"已标记 {added_rows} 行新增（红色）")

    # 保存比较结果文件
    print("\n正在保存结果文件...")
    try:
        wb_baseline.save(output_baseline_path)
        wb_compare.save(output_compare_path)
    except Exception as e:
        print(f"保存结果文件时出错: {e}")
        return
    
    # 生成差异结果文件
    print("\n正在生成差异结果文件...")
    
    # 直接使用保存后的基准文件作为差异结果的基础
    # 这样可以确保格式完全一致
    wb_diff = openpyxl.load_workbook(output_baseline_path)
    ws_diff = wb_diff.active
    ws_diff.title = "差异比较结果"
    
    # 重新加载保存后的文件以获取准确的格式信息
    wb_baseline_saved = openpyxl.load_workbook(output_baseline_path)
    ws_baseline_saved = wb_baseline_saved.active
    
    wb_compare_saved = openpyxl.load_workbook(output_compare_path)
    ws_compare_saved = wb_compare_saved.active
    
    # 创建一个字典来快速查找基准行
    baseline_key_set = set()
    key_to_row = {}
    
    # 获取基准文件中所有行的关键字段值
    data_start_row = header_row + 1
    for row_baseline in range(data_start_row, ws_baseline_saved.max_row + 1):
        if not has_all_keys_baseline:
            continue
        key_values = tuple(ws_baseline_saved.cell(row=row_baseline, column=key_cols_baseline[field]).value for field in key_fields)
        if all(v is not None for v in key_values):
            baseline_key_set.add(key_values)
            key_to_row[key_values] = row_baseline
    
    # 收集比较文件中的新增行（红色行）
    added_rows = []
    for row_compare in range(data_start_row, ws_compare_saved.max_row + 1):
        if not has_all_keys_compare:
            continue
        # 获取当前行的关键字段值
        key_values = tuple(ws_compare_saved.cell(row=row_compare, column=key_cols_compare[field]).value for field in key_fields)
        if not all(v is not None for v in key_values):
            continue
        
        # 检查是否为新增行（红色）
        first_cell = ws_compare_saved.cell(row=row_compare, column=1)
        if first_cell.fill.start_color.rgb == fill_deleted.start_color.rgb:
            # 获取当前行在比较文件中的上一行关键字段值
            prev_key_values = None
            if row_compare > data_start_row:
                prev_key_values = tuple(ws_compare_saved.cell(row=row_compare - 1, column=key_cols_compare[field]).value for field in key_fields)
            added_rows.append((key_values, row_compare, prev_key_values))
    
    # 计算需要插入的行数，提前插入空白行
    for i in range(len(added_rows)):
        # 在差异结果文件末尾插入一行
        ws_diff.append(['' for _ in range(baseline_max_col)])
    
    # 将新增行插入到正确位置
    for key_values, row_compare, prev_key_values in added_rows:
        # 找到插入位置
        insert_row = ws_diff.max_row
        if prev_key_values and prev_key_values in key_to_row:
            insert_row = key_to_row[prev_key_values] + 1
        
        # 插入空白行
        ws_diff.insert_rows(insert_row)
        
        # 更新key_to_row字典
        for k, v in list(key_to_row.items()):
            if v >= insert_row:
                key_to_row[k] = v + 1
        
        # 使用基准文件的数据行作为模板，复制其格式
        template_row = data_start_row
        
        # 先复制模板行的格式到新插入的行
        for col in range(1, baseline_max_col + 1):
            # 获取模板单元格
            template_cell = ws_baseline_saved.cell(row=template_row, column=col)
            new_cell = ws_diff.cell(row=insert_row, column=col)
            
            # 复制数字格式
            new_cell.number_format = template_cell.number_format
            
            # 复制字体样式（创建新的Font对象）
            template_font = template_cell.font
            new_font = Font(
                name=template_font.name,
                size=template_font.size,
                bold=template_font.bold,
                italic=template_font.italic,
                vertAlign=template_font.vertAlign,
                underline=template_font.underline,
                strike=template_font.strike,
                color=template_font.color
            )
            new_cell.font = new_font
            
            # 复制边框样式（创建新的Border对象）
            template_border = template_cell.border
            new_border = Border(
                left=template_border.left,
                right=template_border.right,
                top=template_border.top,
                bottom=template_border.bottom,
                diagonal=template_border.diagonal,
                diagonalUp=template_border.diagonalUp,
                diagonalDown=template_border.diagonalDown
            )
            new_cell.border = new_border
            
            # 复制对齐方式（创建新的Alignment对象）
            template_alignment = template_cell.alignment
            new_alignment = Alignment(
                horizontal=template_alignment.horizontal,
                vertical=template_alignment.vertical,
                textRotation=template_alignment.textRotation,
                wrapText=template_alignment.wrapText,
                shrinkToFit=template_alignment.shrinkToFit,
                indent=template_alignment.indent,
                relativeIndent=template_alignment.relativeIndent,
                justifyLastLine=template_alignment.justifyLastLine,
                readingOrder=template_alignment.readingOrder,
                text_rotation=template_alignment.text_rotation,
                wrap_text=template_alignment.wrap_text,
                shrink_to_fit=template_alignment.shrink_to_fit
            )
            new_cell.alignment = new_alignment
        
        # 然后填入新增行的数据
        for col in range(1, baseline_max_col + 1):
            # 获取基准文件中对应的列名
            col_name_b = ws_baseline_saved.cell(row=header_row, column=col).value
            col_name_b = col_name_b.strip() if col_name_b else ""
            if not col_name_b:
                continue
            
            # 在比较文件中查找对应的列
            for c in range(1, ws_compare_saved.max_column + 1):
                col_name_c = ws_compare_saved.cell(row=header_row, column=c).value
                col_name_c = col_name_c.strip() if col_name_c else ""
                if col_name_c == col_name_b:
                    # 填入数据
                    value = ws_compare_saved.cell(row=row_compare, column=c).value
                    ws_diff.cell(row=insert_row, column=col, value=value)
                    break
        
        # 最后将整行设置为红色填充
        for col in range(1, baseline_max_col + 1):
            cell = ws_diff.cell(row=insert_row, column=col)
            cell.fill = fill_deleted
    
    # 复制基准文件的列宽设置，确保格式完全一致
    for col in range(1, ws_baseline_saved.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in ws_baseline_saved.column_dimensions:
            ws_diff.column_dimensions[col_letter].width = ws_baseline_saved.column_dimensions[col_letter].width
    
    # 复制基准文件的行高设置
    for row in range(1, ws_baseline_saved.max_row + 1):
        if row in ws_baseline_saved.row_dimensions:
            ws_diff.row_dimensions[row].height = ws_baseline_saved.row_dimensions[row].height
    
    # 保存差异结果文件
    # 指向项目根目录的results文件夹
    current_dir = os.path.dirname(os.path.abspath(__file__))
    results_folder = os.path.join(os.path.dirname(current_dir), "results")
    diff_output_path = os.path.join(results_folder, f"{original_filename}_差异结果_{timestamp}.xlsx")
    try:
        wb_diff.save(diff_output_path)
    except Exception as e:
        print(f"保存差异结果文件时出错: {e}")
        return
    
    # 设置文件为只读
    print("\n正在设置文件只读属性...")
    try:
        # 获取当前文件权限
        baseline_stat = os.stat(output_baseline_path)
        compare_stat = os.stat(output_compare_path)
        diff_stat = os.stat(diff_output_path)
        
        # 在Windows上设置只读属性
        if os.name == 'nt':
            # 使用Windows命令设置只读
            subprocess.run(['attrib', '+r', output_baseline_path], check=True, capture_output=True, text=True)
            subprocess.run(['attrib', '+r', output_compare_path], check=True, capture_output=True, text=True)
            subprocess.run(['attrib', '+r', diff_output_path], check=True, capture_output=True, text=True)
        else:
            # 在Linux/macOS上设置只读
            os.chmod(output_baseline_path, baseline_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(output_compare_path, compare_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(diff_output_path, diff_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
        
        print("结果文件已设置为只读属性")
    except Exception as e:
        print(f"设置只读属性时出错: {e}")
    
    # 输出数值变化计数
    if changes_count > 0:
        print(f"已标记 {changes_count} 处数值变化（黄色）")
    
    # 计算总差异数
    total_changes = changes_count
    print(f"\n比较完成！共发现 {total_changes} 处差异。")
    print(f"已生成带颜色标记的文件至: {output_baseline_path}")
    print(f"已生成带颜色标记的文件至: {output_compare_path}")
    print(f"已生成差异结果文件至: {diff_output_path}")


if __name__ == "__main__":
    import sys
    import datetime
    
    if len(sys.argv) != 6:
        print("Usage: python compare_excel_web.py <baseline_path> <compare_path> <output_baseline_path> <output_compare_path> <original_filename>")
        sys.exit(1)
    
    baseline_path = sys.argv[1]
    compare_path = sys.argv[2]
    output_baseline_path = sys.argv[3]
    output_compare_path = sys.argv[4]
    original_filename = sys.argv[5]
    
    # 添加时间戳到输出文件名，避免覆盖现有文件
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 创建results文件夹（如果不存在）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # 指向项目根目录的results文件夹
    results_folder = os.path.join(os.path.dirname(current_dir), "results")
    os.makedirs(results_folder, exist_ok=True)
    
    compare_excel_files(baseline_path, compare_path, output_baseline_path, output_compare_path, original_filename, timestamp)