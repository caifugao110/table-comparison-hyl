import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess
import stat
import sys
import queue
import threading
import datetime
import customtkinter as ctk
from tkinter import filedialog, messagebox
import webbrowser
import io

# 全局队列：用于子线程与GUI线程通信
log_queue = queue.Queue()
progress_queue = queue.Queue()

# 版本和版权信息
VERSION = "V0.0.0"  # 默认版本，会从Gitee动态更新
COPYRIGHT = "Heyanlin © 2026"
PROJECT_URL = "https://github.com/caifugao110/table-comparison-hyl"

# 默认主题设置
DEFAULT_APPEARANCE_MODE = "light"  # "dark", "light", "system"
DEFAULT_COLOR_THEME = "blue"     # "blue", "green", "dark-blue"

# 初始化主题
ctk.set_appearance_mode(DEFAULT_APPEARANCE_MODE)
ctk.set_default_color_theme(DEFAULT_COLOR_THEME)

def compare_excel_files(baseline_path, compare_path, output_baseline_path, output_compare_path, results_folder, original_filename, timestamp, header_row=3, key_fields=None, stop_event=None):
    # 检查停止事件的辅助函数
    def check_stop():
        if stop_event and stop_event.is_set():
            log_queue.put("操作已取消")
            return True
        return False
    
    # 定义颜色样式
    fill_changed = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色：数值变化
    fill_added = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")      # 绿色：新增（在基准基础上）
    fill_deleted = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # 红色：删除（在基准基础上）
    
    try:
        if check_stop():
            return False
            
        # 加载工作簿
        wb_baseline = openpyxl.load_workbook(baseline_path, data_only=True)  # 只加载数据，不加载公式
        wb_compare = openpyxl.load_workbook(compare_path, data_only=True)
    except FileNotFoundError as e:
        log_queue.put(f"错误：找不到文件 - {e}")
        return False
    except Exception as e:
        log_queue.put(f"加载文件时出错: {e}")
        return False

    # 使用第一个工作表
    ws_baseline = wb_baseline.active
    ws_compare = wb_compare.active

    # 获取实际使用的范围
    baseline_max_row = ws_baseline.max_row
    baseline_max_col = ws_baseline.max_column
    compare_max_row = ws_compare.max_row
    compare_max_col = ws_compare.max_column

    # 检查列数是否一致
    if baseline_max_col != compare_max_col:
        log_queue.put(f"警告：两个文件的列数不一致！基准文件：{baseline_max_col}列，比较文件：{compare_max_col}列")

    # 预先获取所有单元格值
    cells_baseline = {}
    cells_compare = {}
    
    # 获取基准文件所有单元格值
    for r in range(1, baseline_max_row + 1):
        if check_stop():
            return False
        for c in range(1, baseline_max_col + 1):
            cells_baseline[(r, c)] = ws_baseline.cell(row=r, column=c).value
    
    # 获取比较文件所有单元格值
    for r in range(1, compare_max_row + 1):
        if check_stop():
            return False
        for c in range(1, compare_max_col + 1):
            cells_compare[(r, c)] = ws_compare.cell(row=r, column=c).value
    
    # 如果没有提供关键字段，默认使用前三列作为特征列
    if not key_fields:
        header_values = [cells_baseline.get((header_row, c), "").strip() for c in range(1, min(baseline_max_col + 1, 4))]
        key_fields = [v for v in header_values if v]  # 过滤空值
        if len(key_fields) < 3:
            key_fields = [f"列{c}" for c in range(1, min(baseline_max_col + 1, 4))]
    
    # 从指定表头行获取关键字段的列索引
    def find_key_columns(cells, max_col, header_row_num, key_field_names):
        """从指定行查找关键字段的列索引"""
        key_cols = {}
        # 获取表头行的所有列名映射
        header_values = {}
        for col in range(1, max_col + 1):
            cell_value = cells.get((header_row_num, col), "").strip()
            header_values[cell_value] = col
        
        # 查找关键字段的列索引
        for field in key_field_names:
            if field in header_values:
                key_cols[field] = header_values[field]
            else:
                # 如果找不到字段名，尝试直接使用列索引
                try:
                    col_idx = int(field.replace("列", ""))
                    if 1 <= col_idx <= max_col:
                        key_cols[field] = col_idx
                except ValueError:
                    pass
        return key_cols
    
    # 查找基准文件和比较文件的关键字段列索引
    key_cols_baseline = find_key_columns(cells_baseline, baseline_max_col, header_row, key_fields)
    key_cols_compare = find_key_columns(cells_compare, compare_max_col, header_row, key_fields)
    
    # 检查是否找到所有关键字段
    has_all_keys_baseline = all(field in key_cols_baseline for field in key_fields)
    has_all_keys_compare = all(field in key_cols_compare for field in key_fields)
    
    # 行匹配：基准行号 -> 比较行号
    row_mapping = {}
    
    if has_all_keys_baseline and has_all_keys_compare:
        # 构建行关键字映射：关键字 -> 行号
        def build_row_key_map(cells, max_row, key_cols, data_start_row):
            row_key_map = {}
            for row in range(data_start_row, max_row + 1):
                key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
                if all(v is not None for v in key_values):
                    row_key_map[key_values] = row
            return row_key_map
        
        # 数据行从表头行的下一行开始
        data_start_row = header_row + 1
        row_key_map_baseline = build_row_key_map(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        row_key_map_compare = build_row_key_map(cells_compare, compare_max_row, key_cols_compare, data_start_row)
        
        # 建立行映射：基准行 -> 比较行
        for key in row_key_map_baseline:
            if key in row_key_map_compare:
                row_baseline = row_key_map_baseline[key]
                row_compare = row_key_map_compare[key]
                row_mapping[row_baseline] = row_compare
    else:
        log_queue.put("\n无法找到所有关键字段，使用默认行匹配...")
        
        # 获取一行的所有单元格内容，作为比较的键
        def get_row_content(row_num, cells, max_col):
            return tuple(cells.get((row_num, c), None) for c in range(1, max_col + 1))
        
        # 构建行内容映射
        row_contents_baseline = {r: get_row_content(r, cells_baseline, baseline_max_col) for r in range(1, baseline_max_row + 1)}
        row_contents_compare = {r: get_row_content(r, cells_compare, compare_max_col) for r in range(1, compare_max_row + 1)}
        
        # 先找到完全匹配的行
        for row_baseline, content_baseline in row_contents_baseline.items():
            if check_stop():
                return False
            
            for row_compare, content_compare in row_contents_compare.items():
                if row_compare not in row_mapping.values() and content_baseline == content_compare:
                    row_mapping[row_baseline] = row_compare
                    break
        
        # 如果没有找到足够的匹配，使用简单的索引映射
        if len(row_mapping) < min(baseline_max_row, compare_max_row) // 2:
            min_rows = min(baseline_max_row, compare_max_row)
            row_mapping = {r: r for r in range(1, min_rows + 1)}
    
    # 比较单元格并标记差异
    changes_count = 0  # 数值变化计数
    added_rows_count = 0  # 新增行计数
    deleted_rows_count = 0  # 删除行计数
    
    # 定义关键字段列索引集合，避免重新计算
    key_col_set_baseline = set(key_cols_baseline.values()) if has_all_keys_baseline else set()
    key_col_set_compare = set(key_cols_compare.values()) if has_all_keys_compare else set()
    
    log_queue.put("\n开始比较匹配行的单元格差异...")
    
    # 创建列映射（基于列名匹配）
    def create_col_name_map():
        col_name_map = {}
        # 先获取基准文件的列名映射
        baseline_col_names = {}
        for col_b in range(1, baseline_max_col + 1):
            col_name_b = cells_baseline.get((header_row, col_b), "").strip()
            if col_name_b:
                baseline_col_names[col_name_b] = col_b
        
        # 然后在比较文件中查找相同列名
        for col_c in range(1, compare_max_col + 1):
            col_name_c = cells_compare.get((header_row, col_c), "").strip()
            if col_name_c in baseline_col_names:
                col_name_map[baseline_col_names[col_name_c]] = col_c
        
        # 如果没有找到足够的匹配，使用简单的索引映射
        if len(col_name_map) < min(baseline_max_col, compare_max_col) // 2:
            min_cols = min(baseline_max_col, compare_max_col)
            col_name_map = {c: c for c in range(1, min_cols + 1)}
        
        return col_name_map
    
    col_name_map = create_col_name_map()
    
    for row_baseline, row_compare in row_mapping.items():
        if check_stop():
            return False
            
        # 比较匹配的列
        for col_baseline, col_compare in col_name_map.items():
            # 跳过关键字段列（它们已经匹配，不需要比较）
            if col_baseline in key_col_set_baseline or col_compare in key_col_set_compare:
                continue
            
            val_baseline = cells_baseline.get((row_baseline, col_baseline), None)
            val_compare = cells_compare.get((row_compare, col_compare), None)
            
            # 只在值不同时标记为黄色（数值变化）
            if val_baseline != val_compare:
                ws_baseline.cell(row=row_baseline, column=col_baseline).fill = fill_changed
                ws_compare.cell(row=row_compare, column=col_compare).fill = fill_changed
                changes_count += 1
    
    log_queue.put("\n开始标记新增行、删除行和数值变化行...")
    
    # 获取所有数据行的关键字映射
    def get_all_row_keys(cells, max_row, key_cols, data_start_row):
        """获取所有数据行的关键字映射"""
        all_row_keys = {}
        for row in range(data_start_row, max_row + 1):
            key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
            if all(v is not None for v in key_values):
                all_row_keys[key_values] = row
        return all_row_keys
    
    if has_all_keys_baseline and has_all_keys_compare:
        # 获取所有数据行的关键字映射
        data_start_row = header_row + 1
        all_baseline_keys = get_all_row_keys(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        all_compare_keys = get_all_row_keys(cells_compare, compare_max_row, key_cols_compare, data_start_row)
        
        # 标记删除行（基准文件中有，比较文件中没有）
        for key, row_baseline in all_baseline_keys.items():
            if check_stop():
                return False
                
            if key not in all_compare_keys:
                # 标记整行为绿色
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                deleted_rows_count += 1
        log_queue.put(f"\n已标记 {deleted_rows_count} 行删除（绿色）")
        
        # 标记新增行（比较文件中有，基准文件中没有）
        for key, row_compare in all_compare_keys.items():
            if check_stop():
                return False
                
            if key not in all_baseline_keys:
                # 标记整行为红色
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                added_rows_count += 1
        log_queue.put(f"\n已标记 {added_rows_count} 行新增（红色）")
    else:
        # 使用简单的行匹配来标记新增和删除行
        log_queue.put("\n使用简单匹配标记新增和删除行...")
        
        # 标记删除行（基准文件中有，比较文件中没有对应的行）
        for row_baseline in range(1, baseline_max_row + 1):
            if check_stop():
                return False
                
            if row_baseline not in row_mapping:
                # 标记整行为绿色
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                deleted_rows_count += 1
        log_queue.put(f"\n已标记 {deleted_rows_count} 行删除（绿色）")
        
        # 标记新增行（比较文件中有，基准文件中没有对应的行）
        mapped_compare_rows = set(row_mapping.values())
        for row_compare in range(1, compare_max_row + 1):
            if check_stop():
                return False
                
            if row_compare not in mapped_compare_rows:
                # 标记整行为红色
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                added_rows_count += 1
        log_queue.put(f"\n已标记 {added_rows_count} 行新增（红色）")
    
    # 输出数值变化行计数
    if changes_count > 0:
        log_queue.put(f"\n已标记 {changes_count} 处数值变化（黄色）")

    # 计算总差异数
    total_changes = changes_count + added_rows_count + deleted_rows_count
    log_queue.put(f"\n比较完成！共发现 {total_changes} 处差异。")

    # 保存比较结果文件
    try:
        wb_baseline.save(output_baseline_path)
        wb_compare.save(output_compare_path)
    except Exception as e:
        log_queue.put(f"保存结果文件时出错: {e}")
        return False
    
    # 生成差异结果文件
    log_queue.put("\n正在生成差异结果文件...")
    
    try:
        # 使用保存后的基准文件作为差异结果的基础
        wb_diff = openpyxl.load_workbook(output_baseline_path)
        ws_diff = wb_diff.active
        ws_diff.title = "差异比较结果"
        
        # 重新加载保存后的文件以获取准确的格式信息
        wb_baseline_saved = openpyxl.load_workbook(output_baseline_path)
        ws_baseline_saved = wb_baseline_saved.active
        
        wb_compare_saved = openpyxl.load_workbook(output_compare_path)
        ws_compare_saved = wb_compare_saved.active
    except Exception as e:
        log_queue.put(f"加载保存后的文件时出错: {e}")
        return False
    
    # 创建一个字典来快速查找基准行
    key_to_row = {}
    
    # 获取基准文件中所有行的关键字段值
    for row_baseline in range(4, ws_baseline_saved.max_row + 1):
        if check_stop():
            return False
            
        key_values = tuple(ws_baseline_saved.cell(row=row_baseline, column=key_cols_baseline[field]).value for field in key_fields)
        if all(v is not None for v in key_values):
            key_to_row[key_values] = row_baseline
    
    # 收集比较文件中的新增行（红色行）
    added_rows = []
    for row_compare in range(4, ws_compare_saved.max_row + 1):
        if check_stop():
            return False
            
        # 获取当前行的关键字段值
        key_values = tuple(ws_compare_saved.cell(row=row_compare, column=key_cols_compare[field]).value for field in key_fields)
        if not all(v is not None for v in key_values):
            continue
        
        # 检查是否为新增行（红色）
        first_cell = ws_compare_saved.cell(row=row_compare, column=1)
        if first_cell.fill.start_color.rgb == fill_deleted.start_color.rgb:
            # 获取当前行在比较文件中的上一行关键字段值
            prev_key_values = None
            if row_compare > 4:
                prev_key_values = tuple(ws_compare_saved.cell(row=row_compare - 1, column=key_cols_compare[field]).value for field in key_fields)
            added_rows.append((key_values, row_compare, prev_key_values))
    
    # 计算需要插入的行数，提前插入空白行
    for _ in range(len(added_rows)):
        if check_stop():
            return False
        ws_diff.append(['' for _ in range(baseline_max_col)])
    
    # 将新增行插入到正确位置
    for key_values, row_compare, prev_key_values in added_rows:
        if check_stop():
            return False
            
        # 找到插入位置
        insert_row = ws_diff.max_row
        if prev_key_values and prev_key_values in key_to_row:
            insert_row = key_to_row[prev_key_values] + 1
        
        # 插入空白行
        ws_diff.insert_rows(insert_row)
        
        # 更新key_to_row字典
        for k, v in list(key_to_row.items()):
            if v >= insert_row:
                key_to_row[k] = v + 1
        
        # 使用基准文件的第4行作为模板，复制其格式
        template_row = 4
        
        # 先复制模板行的格式到新插入的行
        for col in range(1, baseline_max_col + 1):
            template_cell = ws_baseline_saved.cell(row=template_row, column=col)
            new_cell = ws_diff.cell(row=insert_row, column=col)
            
            # 复制格式
            new_cell.number_format = template_cell.number_format
            new_cell.font = Font(**template_cell.font.__dict__)
            new_cell.border = Border(**template_cell.border.__dict__)
            new_cell.alignment = Alignment(**template_cell.alignment.__dict__)
        
        # 然后填入新增行的数据
        for col in range(1, baseline_max_col + 1):
            # 获取基准文件中对应的列名
            col_name_b = ws_baseline_saved.cell(row=3, column=col).value
            col_name_b = col_name_b.strip() if col_name_b else ""
            if not col_name_b:
                continue
            
            # 在比较文件中查找对应的列
            for c in range(1, ws_compare_saved.max_column + 1):
                col_name_c = ws_compare_saved.cell(row=3, column=c).value
                col_name_c = col_name_c.strip() if col_name_c else ""
                if col_name_c == col_name_b:
                    # 填入数据
                    value = ws_compare_saved.cell(row=row_compare, column=c).value
                    ws_diff.cell(row=insert_row, column=col, value=value)
                    break
        
        # 最后将整行设置为红色填充
        for col in range(1, baseline_max_col + 1):
            ws_diff.cell(row=insert_row, column=col).fill = fill_deleted
    
    # 复制基准文件的列宽设置
    for col in range(1, ws_baseline_saved.max_column + 1):
        if check_stop():
            return False
            
        col_letter = get_column_letter(col)
        if col_letter in ws_baseline_saved.column_dimensions:
            ws_diff.column_dimensions[col_letter].width = ws_baseline_saved.column_dimensions[col_letter].width
    
    # 复制基准文件的行高设置
    for row in range(1, ws_baseline_saved.max_row + 1):
        if check_stop():
            return False
            
        if row in ws_baseline_saved.row_dimensions:
            ws_diff.row_dimensions[row].height = ws_baseline_saved.row_dimensions[row].height
    
    # 保存差异结果文件
    diff_output_path = os.path.join(results_folder, f"{original_filename}_差异结果_{timestamp}.xlsx")
    try:
        wb_diff.save(diff_output_path)
    except Exception as e:
        log_queue.put(f"保存差异结果文件时出错: {e}")
        return False
    
    # 设置文件为只读
    try:
        # 获取当前文件权限
        baseline_stat = os.stat(output_baseline_path)
        compare_stat = os.stat(output_compare_path)
        diff_stat = os.stat(diff_output_path)
        
        # 在Windows上设置只读属性
        if os.name == 'nt':
            subprocess.run(['attrib', '+r', output_baseline_path], check=True)
            subprocess.run(['attrib', '+r', output_compare_path], check=True)
            subprocess.run(['attrib', '+r', diff_output_path], check=True)
        else:
            # 在Linux/macOS上设置只读
            os.chmod(output_baseline_path, baseline_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(output_compare_path, compare_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(diff_output_path, diff_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
    except Exception as e:
        log_queue.put(f"设置只读属性时出错: {e}")
    
    log_queue.put(f"\n已生成差异结果文件至: \n{diff_output_path}")
    
    # 自动打开文件
    try:
        subprocess.Popen(['start', '', output_baseline_path], shell=True)
        subprocess.Popen(['start', '', output_compare_path], shell=True)
        subprocess.Popen(['start', '', diff_output_path], shell=True)
    except Exception as e:
        log_queue.put(f"打开文件时出错: {e}")
    
    return True



class StdoutRedirector:
    """重定向stdout到GUI的Text组件"""
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        log_queue.put(message)

    def flush(self):
        pass

class ExcelCompareGUI(ctk.CTk):
    """Excel文件比较工具GUI界面"""
    def __init__(self):
        super().__init__()
        self.title("Excel文件比较工具")
        self.geometry("1200x800")
        self.minsize(1000, 700)
        
        # 设置窗口图标
        self.app_icon = None
        try:
            # 使用PIL创建图标
            from PIL import Image, ImageDraw, ImageFont, ImageTk
            
            # 创建一个32x32的图标
            icon = Image.new('RGB', (32, 32), color=(102, 126, 234))
            draw = ImageDraw.Draw(icon)
            
            # 绘制Excel表格和比较相关的图形
            draw.rectangle([4, 8, 14, 24], fill='white', outline='white')
            draw.rectangle([18, 8, 28, 24], fill='white', outline='white')
            
            # 在矩形上绘制比较符号
            draw.text((8, 12), 'A', fill=(102, 126, 234), font=ImageFont.truetype('arial.ttf', 10))
            draw.text((22, 12), 'B', fill=(102, 126, 234), font=ImageFont.truetype('arial.ttf', 10))
            
            # 绘制比较箭头
            draw.line([15, 16, 18, 16], fill='white', width=2)
            draw.polygon([18, 14, 18, 18, 21, 16], fill='white')
            
            # 保存图标以便后续使用
            self.app_icon = ImageTk.PhotoImage(icon)
            self.iconphoto(False, self.app_icon)
        except Exception as e:
            print(f"设置图标失败: {e}")
        
        # 配置变量
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.parent_dir = os.path.dirname(self.current_dir)
        self.results_folder = os.path.join(self.parent_dir, "results")
        os.makedirs(self.results_folder, exist_ok=True)
        
        self.baseline_file = ""
        self.compare_file = ""
        self.running = False
        self.stop_event = threading.Event()
        self.worker_thread = None
        
        # 初始化界面
        self._init_widgets()
        
        # 动态获取版本信息
        self._get_latest_version()
        
        # 重定向stdout
        self._redirect_stdout()
        
        # 启动队列监听
        self._listen_queues()
    
    def _init_widgets(self):
        """初始化GUI组件"""
        # 创建主容器
        main_container = ctk.CTkFrame(self)
        main_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        # 顶部标题栏
        header_frame = ctk.CTkFrame(main_container, fg_color=("gray90", "gray20"), height=100)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # 标题和主题选择
        title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=10)
        
        # 标题
        title_label = ctk.CTkLabel(
            title_frame, 
            text="Excel文件比较工具",
            font=("微软雅黑", 26, "bold"),
            text_color=("#1f77b4", "#64b5f6")
        )
        title_label.pack(anchor="w", side="left")
        
        # 主题选择
        theme_frame = ctk.CTkFrame(title_frame, fg_color="transparent")
        theme_frame.pack(anchor="e", side="right")
        
        ctk.CTkLabel(
            theme_frame,
            text="主题:",
            font=("微软雅黑", 12),
            text_color=("gray50", "gray70")
        ).pack(side="left", padx=(0, 10))
        
        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(
            theme_frame,
            values=["light", "dark", "system"],
            command=self._change_appearance_mode_event,
            font=("微软雅黑", 12),
            width=120
        )
        self.appearance_mode_optionemenu.set(DEFAULT_APPEARANCE_MODE)
        self.appearance_mode_optionemenu.pack(side="left", padx=(0, 10))
        
        self.color_theme_optionemenu = ctk.CTkOptionMenu(
            theme_frame,
            values=["blue", "green", "dark-blue"],
            command=self._change_color_theme_event,
            font=("微软雅黑", 12),
            width=120
        )
        self.color_theme_optionemenu.set(DEFAULT_COLOR_THEME)
        self.color_theme_optionemenu.pack(side="left")
        
        # 版本和链接信息
        info_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        info_frame.pack(anchor="w", padx=20, pady=(0, 10))
        
        self.version_label = ctk.CTkLabel(
            info_frame,
            text=f"{COPYRIGHT} | {VERSION}",
            font=("微软雅黑", 12),
            text_color=("gray50", "gray70")
        )
        self.version_label.pack(side="left", padx=(0, 20))
        
        github_btn = ctk.CTkButton(
            info_frame,
            text="📌 GitHub地址",
            width=120,
            height=30,
            font=("微软雅黑", 12),
            command=lambda: webbrowser.open(PROJECT_URL)
        )
        github_btn.pack(side="left", padx=5)
        
        help_btn = ctk.CTkButton(
            info_frame,
            text="❓ 使用说明",
            width=120,
            height=30,
            font=("微软雅黑", 12),
            command=lambda: webbrowser.open("https://github.com/caifugao110/table-comparison-hyl/blob/master/README.md")
        )
        help_btn.pack(side="left", padx=5)
        
        # 主内容区
        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # 左侧面板（文件选择和操作）
        left_panel = ctk.CTkFrame(content_frame, fg_color=("gray86", "gray17"))
        left_panel.pack(side="left", fill="y", expand=False, padx=(0, 10))
        left_panel.configure(width=300)
        
        # 文件选择区
        file_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        file_section.pack(fill="x", padx=15, pady=15)
        
        ctk.CTkLabel(
            file_section, 
            text="文件选择", 
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w", pady=(0, 10))
        
        # 基准文件选择
        baseline_frame = ctk.CTkFrame(file_section, fg_color="transparent")
        baseline_frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            baseline_frame, 
            text="基准文件:", 
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")
        
        self.baseline_entry = ctk.CTkEntry(baseline_frame, font=("微软雅黑", 12))
        self.baseline_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        ctk.CTkButton(
            baseline_frame, 
            text="浏览", 
            width=60,
            font=("微软雅黑", 12),
            command=self._browse_baseline_file
        ).pack(side="left", padx=5)
        
        # 比较文件选择
        compare_frame = ctk.CTkFrame(file_section, fg_color="transparent")
        compare_frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            compare_frame, 
            text="比较文件:", 
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")
        
        self.compare_entry = ctk.CTkEntry(compare_frame, font=("微软雅黑", 12))
        self.compare_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        ctk.CTkButton(
            compare_frame, 
            text="浏览", 
            width=60,
            font=("微软雅黑", 12),
            command=self._browse_compare_file
        ).pack(side="left", padx=5)
        
        # 配置选项区
        config_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        config_section.pack(fill="x", padx=15, pady=15)
        
        ctk.CTkLabel(
            config_section, 
            text="比较配置", 
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w", pady=(0, 10))
        
        # 表头行号选择
        header_row_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        header_row_frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            header_row_frame, 
            text="表头行号:", 
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")
        
        self.header_row_var = ctk.StringVar(value="")
        self.header_row_entry = ctk.CTkEntry(header_row_frame, textvariable=self.header_row_var, font=("微软雅黑", 12), width=150, state="readonly")
        self.header_row_entry.pack(side="left", padx=5)
        
        ctk.CTkButton(
            header_row_frame, 
            text="选择", 
            width=60,
            font=("微软雅黑", 12),
            command=self._select_header_row
        ).pack(side="left", padx=5)
        
        # 表头行预览信息
        self.header_preview_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        self.header_preview_frame.pack(fill="x", pady=5)
        
        self.header_preview_label = ctk.CTkLabel(
            self.header_preview_frame, 
            text="请点击'选择'按钮查看并选择表头行号", 
            font=("微软雅黑", 10),
            text_color="gray50"
        )
        self.header_preview_label.pack(anchor="w")
        
        # 特征列选择
        feature_cols_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        feature_cols_frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            feature_cols_frame, 
            text="特征列:", 
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")
        
        self.feature_cols_var = ctk.StringVar(value="1,2,3")
        self.feature_cols_entry = ctk.CTkEntry(feature_cols_frame, textvariable=self.feature_cols_var, font=("微软雅黑", 12), width=150, state="readonly")
        self.feature_cols_entry.pack(side="left", padx=5)
        
        ctk.CTkButton(
            feature_cols_frame, 
            text="选择", 
            width=60,
            font=("微软雅黑", 12),
            command=self._select_feature_columns
        ).pack(side="left", padx=5)
        
        # 特征列预览信息
        self.feature_cols_preview_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        self.feature_cols_preview_frame.pack(fill="x", pady=5)
        
        self.feature_cols_preview_label = ctk.CTkLabel(
            self.feature_cols_preview_frame, 
            text="请点击'选择'按钮查看并选择特征列，最多支持6列，默认使用列: 1,2,3", 
            font=("微软雅黑", 10),
            text_color="gray50"
        )
        self.feature_cols_preview_label.pack(anchor="w")
        
        ctk.CTkLabel(
            config_section, 
            text="提示: 特征列用于判断行的增删变化，特征列内容的变化不视为数值变化", 
            font=("微软雅黑", 12, "bold"),
            text_color="#FF6B35"
        ).pack(anchor="w", pady=(5, 0))
        
        # 操作按钮区
        button_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        button_section.pack(fill="x", padx=15, pady=15)
        
        self.start_button = ctk.CTkButton(
            button_section, 
            text="开始比较", 
            font=("微软雅黑", 16, "bold"),
            height=50,
            fg_color="#4CAF50",
            hover_color="#45a049",
            command=self._start_compare
        )
        self.start_button.pack(fill="x", pady=5)
        
        self.stop_button = ctk.CTkButton(
            button_section, 
            text="停止", 
            font=("微软雅黑", 16, "bold"),
            height=50,
            fg_color="#f44336",
            hover_color="#da190b",
            command=self._stop_compare,
            state="disabled"
        )
        self.stop_button.pack(fill="x", pady=5)
        
        # 右侧面板（日志显示）
        right_panel = ctk.CTkFrame(content_frame, fg_color=("gray86", "gray17"))
        right_panel.pack(side="right", fill="both", expand=True)
        
        # 日志标题
        log_title_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        log_title_frame.pack(fill="x", padx=15, pady=10)
        
        ctk.CTkLabel(
            log_title_frame, 
            text="任务日志", 
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w")
        
        # 日志显示区域
        log_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        log_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=("微软雅黑", 12),
            wrap="word",
            corner_radius=8,
            border_width=2,
            border_color=("#D1D1D6", "#4A4A4A"),
            fg_color=("#F8F8F8", "#1A1A1A"),
            text_color=("#424242", "#B0BEC5"),
            padx=10,
            pady=10,
            height=80  # 进一步减小高度
        )
        # 使用grid布局替代pack，更好地控制滚动条
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # 滚动条
        scrollbar = ctk.CTkScrollbar(
            log_frame,
            command=self.log_text.yview,
            corner_radius=8
        )
        scrollbar.grid(row=0, column=1, sticky="ns", padx=(0, 5), pady=5)
        self.log_text.configure(yscrollcommand=scrollbar.set)
    
    def _change_appearance_mode_event(self, new_appearance_mode: str):
        """切换外观模式"""
        ctk.set_appearance_mode(new_appearance_mode)
    
    def _change_color_theme_event(self, new_color_theme: str):
        """切换颜色主题"""
        ctk.set_default_color_theme(new_color_theme)
    
    def _browse_baseline_file(self):
        """浏览基准文件"""
        file_path = filedialog.askopenfilename(
            title="选择基准Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*")]
        )
        if file_path:
            self.baseline_entry.delete(0, ctk.END)
            self.baseline_entry.insert(0, file_path)
            self.baseline_file = file_path
    
    def _browse_compare_file(self):
        """浏览比较文件"""
        file_path = filedialog.askopenfilename(
            title="选择比较Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*")]
        )
        if file_path:
            self.compare_entry.delete(0, ctk.END)
            self.compare_entry.insert(0, file_path)
            self.compare_file = file_path
    
    def _start_compare(self):
        """开始比较"""
        # 检查文件是否选择
        self.baseline_file = self.baseline_entry.get().strip()
        self.compare_file = self.compare_entry.get().strip()
        
        if not self.baseline_file or not self.compare_file:
            messagebox.showerror("错误", "请选择基准文件和比较文件")
            return
        
        if not os.path.exists(self.baseline_file):
            messagebox.showerror("错误", f"基准文件不存在: {self.baseline_file}")
            return
        
        if not os.path.exists(self.compare_file):
            messagebox.showerror("错误", f"比较文件不存在: {self.compare_file}")
            return
        
        # 检查表头行号是否已选择
        header_row_str = self.header_row_var.get().strip()
        if not header_row_str:
            messagebox.showerror("错误", "请选择表头行号")
            return
        
        # 开始比较
        self.running = True
        self.stop_event.clear()
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        
        # 清空日志
        self.log_text.delete("1.0", ctk.END)
        
        # 创建工作线程
        self.worker_thread = threading.Thread(
            target=self._compare_worker,
            daemon=True
        )
        self.worker_thread.start()
    
    def _stop_compare(self):
        """停止比较"""
        self.stop_event.set()
        self.stop_button.configure(state="disabled")
    
    def _select_header_row(self):
        """选择表头行号"""
        if not self.baseline_file:
            messagebox.showerror("错误", "请先选择基准文件")
            return
        
        try:
            # 加载基准文件获取数据
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active
            
            # 获取前10行数据
            max_row = min(10, ws.max_row)
            max_col = min(6, ws.max_column)
            
            # 创建表头行选择窗口
            select_window = ctk.CTkToplevel(self)
            select_window.title("选择表头行号")
            select_window.geometry("900x400")
            
            # 设置窗口图标与主窗口一致
            if hasattr(self, 'app_icon') and self.app_icon:
                select_window.iconphoto(False, self.app_icon)
            
            # 居中显示
            select_window.transient(self)
            select_window.grab_set()
            
            # 创建表格预览区
            preview_frame = ctk.CTkFrame(select_window)
            preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 显示行号和数据
            for row in range(1, max_row + 1):
                # 行号按钮
                row_btn = ctk.CTkButton(
                    preview_frame, 
                    text=f"行 {row}", 
                    width=60,
                    height=30,
                    font=("微软雅黑", 10),
                    command=lambda r=row: self._set_header_row(r, select_window)
                )
                row_btn.grid(row=row, column=0, padx=5, pady=2, sticky="w")
                
                # 显示前6列数据
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    cell_text = str(cell_value) if cell_value else "空"
                    
                    cell_label = ctk.CTkLabel(
                        preview_frame, 
                        text=cell_text,
                        width=140,
                        height=30,
                        font=("微软雅黑", 10),
                        anchor="w"
                    )
                    cell_label.grid(row=row, column=col, padx=5, pady=2, sticky="w")
            
            # 说明文字
            info_label = ctk.CTkLabel(
                select_window, 
                text="请点击行号选择表头所在行", 
                font=("微软雅黑", 12)
            )
            info_label.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
    
    def _set_header_row(self, row_num, window):
        """设置表头行号并关闭窗口"""
        self.header_row_var.set(str(row_num))
        
        # 更新预览信息
        try:
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active
            
            # 获取所选行的前6列数据
            cols_data = []
            max_col = min(6, ws.max_column)
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row_num, column=col).value
                cols_data.append(f"列{col}={str(cell_value) if cell_value else '空'}")
            
            self.header_preview_label.configure(
                text=f"已选择表头行 {row_num}，内容预览: {', '.join(cols_data)}"
            )
        except Exception as e:
            self.header_preview_label.configure(
                text=f"已选择表头行 {row_num}"
            )
        
        window.destroy()
    
    def _select_feature_columns(self):
        """选择特征列"""
        if not self.baseline_file:
            messagebox.showerror("错误", "请先选择基准文件")
            return
        
        try:
            # 加载基准文件获取表头信息
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active
            
            # 获取用户选择的表头行号
            try:
                header_row = int(self.header_row_var.get())
            except ValueError:
                messagebox.showerror("错误", "表头行号必须是数字")
                return
            
            # 获取表头行的列名
            max_col = ws.max_column
            header_values = []
            # 创建列号到列名的映射字典
            col_name_map = {}
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                col_name = cell_value.strip() if cell_value else "空"
                header_values.append(f"{col}: {col_name}")
                col_name_map[col] = col_name
            
            # 创建特征列选择窗口
            select_window = ctk.CTkToplevel(self)
            select_window.title("选择特征列")
            select_window.geometry("400x300")
            select_window.resizable(False, False)
            
            # 设置窗口图标与主窗口一致
            if hasattr(self, 'app_icon') and self.app_icon:
                select_window.iconphoto(False, self.app_icon)
            
            # 居中显示
            select_window.transient(self)
            select_window.grab_set()
            
            # 创建列表框
            listbox = ctk.CTkScrollableFrame(select_window)
            listbox.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 创建复选框
            checkboxes = []
            for i, header in enumerate(header_values[:20]):  # 最多显示20列
                var = ctk.IntVar()
                checkbox = ctk.CTkCheckBox(listbox, text=header, variable=var)
                checkbox.pack(anchor="w", pady=5)
                checkboxes.append((var, i + 1))  # 列号从1开始
            
            # 选择按钮
            def on_select():
                selected = [col for var, col in checkboxes if var.get() == 1]
                if len(selected) == 0:
                    messagebox.showerror("错误", "请至少选择1列")
                    return
                if len(selected) > 6:
                    messagebox.showerror("错误", "最多只能选择6列")
                    return
                
                # 更新特征列显示
                selected_str = ", ".join(map(str, selected))
                self.feature_cols_var.set(selected_str)
                
                # 显示列名预览
                selected_col_names = [f"{col}({col_name_map[col]})" for col in selected]
                preview_text = f"已选择特征列: {', '.join(selected_col_names)}"
                
                # 更新预览信息
                self.feature_cols_preview_label.configure(
                    text=preview_text
                )
                
                select_window.destroy()
            
            select_button = ctk.CTkButton(select_window, text="确定", command=on_select, fg_color="#4CAF50")
            select_button.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
    
    def _compare_worker(self):
        """比较工作线程"""
        try:
            # 输出已选择的文件
            log_queue.put(f"已选择基准文件 {self.baseline_file}")
            log_queue.put(f"\n已选择比较文件 {self.compare_file}")
            
            # 获取表头行号
            try:
                header_row = int(self.header_row_var.get())
            except ValueError:
                log_queue.put("\n❌ 错误：表头行号必须是数字")
                return False
            
            # 获取特征列
            feature_cols_str = self.feature_cols_var.get()
            key_fields = None
            try:
                # 解析特征列，支持多种格式："1,2,3" 或 "1 2 3" 或 "1-3"
                feature_cols = []
                # 处理逗号分隔
                parts = [p.strip() for p in feature_cols_str.split(",")]
                for part in parts:
                    # 处理空格分隔
                    sub_parts = [sp.strip() for sp in part.split() if sp.strip()]
                    for sub_part in sub_parts:
                        # 处理范围
                        if "-" in sub_part:
                            start, end = map(int, sub_part.split("-"))
                            feature_cols.extend(range(start, end + 1))
                        else:
                            feature_cols.append(int(sub_part))
                # 去重并排序
                feature_cols = sorted(list(set(feature_cols)))
                # 转换为列名格式
                key_fields = [f"列{col}" for col in feature_cols]
            except ValueError:
                log_queue.put("\n❌ 错误：特征列格式无效")
                return False
            
            # 生成结果文件名
            baseline_folder = os.path.basename(os.path.dirname(self.baseline_file))
            compare_folder = os.path.basename(os.path.dirname(self.compare_file))
            original_filename = os.path.basename(self.baseline_file).replace('.xlsx', '')
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 读取表头行内容用于预览
            header_preview = ""
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
                ws = wb.active
                if header_row <= ws.max_row:
                    # 获取表头行的前6列内容作为预览
                    max_col = min(6, ws.max_column)
                    header_cells = []
                    for col in range(1, max_col + 1):
                        cell_value = ws.cell(row=header_row, column=col).value
                        if cell_value:
                            header_cells.append(str(cell_value))
                        else:
                            header_cells.append("空")
                    header_preview = ", ".join(header_cells)
                    if ws.max_column > 6:
                        header_preview += f", ... (共{ws.max_column}列)"
            except Exception as e:
                header_preview = "无法读取表头内容"
            
            # 输出比较配置信息
            log_queue.put("\n已定义比较配置：")
            log_queue.put(f"\n已选择表头行 {header_row}")
            log_queue.put(f"\n已选择特征列：{feature_cols_str}")
            
            # 构建结果文件路径
            result_baseline = os.path.join(
                self.results_folder, 
                f"{original_filename}_{baseline_folder}_比较结果_{timestamp}.xlsx"
            )
            result_compare = os.path.join(
                self.results_folder, 
                f"{original_filename}_{compare_folder}_比较结果_{timestamp}.xlsx"
            )
            
            # 调用比较函数
            success = compare_excel_files(
                self.baseline_file, 
                self.compare_file, 
                result_baseline, 
                result_compare,
                self.results_folder,
                original_filename,
                timestamp,
                header_row,
                key_fields,
                self.stop_event
            )
            
            if success:
                log_queue.put("\n✅ 任务完成！")
            else:
                log_queue.put("\n❌ 任务失败！")
        except Exception as e:
            log_queue.put(f"\n❌ 任务过程中出错: {str(e)}")
        finally:
            # 更新UI状态
            self.running = False
            self.start_button.configure(state="normal")
            self.stop_button.configure(state="disabled")
    
    def _redirect_stdout(self):
        """重定向标准输出到日志组件"""
        sys.stdout = StdoutRedirector(self.log_text)
    
    def _get_latest_version(self):
        """从Gitee获取最新版本信息"""
        import threading
        import requests
        
        def fetch_version():
            """在后台线程中获取版本信息"""
            try:
                # Gitee API参数
                owner = "caifugao110"
                repo = "table-comparison-hyl"
                gitee_token = "a09da64c1d9e9c7420a18dfd838890b0"
                headers = {
                    "Authorization": f"token {gitee_token}",
                    "Accept": "application/json"
                }
                
                # 全局变量，用于更新
                global VERSION
                
                # 获取最新发行版本
                try:
                    release_url = f"https://gitee.com/api/v5/repos/{owner}/{repo}/releases/latest"
                    release_response = requests.get(release_url, headers=headers, timeout=10)
                    if release_response.status_code == 200:
                        release_data = release_response.json()
                        latest_version = release_data.get("tag_name")
                        if latest_version:
                            VERSION = latest_version
                except Exception as e:
                    # 静默处理，不打印调试信息
                    pass
                
                # 更新UI中的版本信息
                self.after(0, self._update_version_info)
            except Exception as e:
                print(f"获取版本信息失败: {e}")
        
        # 在后台线程中执行，避免阻塞GUI
        thread = threading.Thread(target=fetch_version, daemon=True)
        thread.start()
    
    def _update_version_info(self):
        """更新UI中的版本信息"""
        # 直接更新版本信息标签
        new_text = f"{COPYRIGHT} | {VERSION}"
        self.version_label.configure(text=new_text)
    
    def _listen_queues(self):
        """监听日志队列并更新UI"""
        try:
            while not log_queue.empty():
                message = log_queue.get_nowait()
                # 确保每条日志单独一行
                if not message.endswith('\n'):
                    message += '\n'
                
                # 插入日志
                self.log_text.insert(ctk.END, message)
                
                # 简化颜色方案，统一使用深色主题下的白色和浅色主题下的黑色作为基础颜色
                line_start = "end-2l"
                line_end = "end-1l"
                
                # 根据日志内容设置不同颜色
                if "错误" in message or "Error" in message or "ERROR" in message or "出错" in message:
                    # 错误日志 - 红色
                    self.log_text.tag_add("error", line_start, line_end)
                    self.log_text.tag_config("error", foreground="#FF5252")
                elif "警告" in message or "Warning" in message or "WARNING" in message:
                    # 警告日志 - 橙色
                    self.log_text.tag_add("warning", line_start, line_end)
                    self.log_text.tag_config("warning", foreground="#FF9800")
                elif "取消" in message:
                    # 取消日志 - 灰色
                    self.log_text.tag_add("cancel", line_start, line_end)
                    self.log_text.tag_config("cancel", foreground="#9E9E9E")
                elif "完成" in message or "成功" in message or "完成!" in message:
                    # 成功日志 - 绿色
                    self.log_text.tag_add("success", line_start, line_end)
                    self.log_text.tag_config("success", foreground="#4CAF50")
                elif "开始" in message or "正在" in message:
                    # 进程日志 - 蓝色
                    self.log_text.tag_add("process", line_start, line_end)
                    self.log_text.tag_config("process", foreground="#2196F3")
                elif "已标记" in message or "共发现" in message or "生成" in message:
                    # 结果日志 - 紫色
                    self.log_text.tag_add("result", line_start, line_end)
                    self.log_text.tag_config("result", foreground="#9C27B0")
                else:
                    # 普通日志 - 黑色/白色
                    self.log_text.tag_add("normal", line_start, line_end)
                    self.log_text.tag_config("normal", foreground="#424242")
                
                self.log_text.see(ctk.END)
        except queue.Empty:
            pass
        finally:
            # 每100ms检查一次队列
            self.after(100, self._listen_queues)

if __name__ == "__main__":
    app = ExcelCompareGUI()
    app.mainloop()